import streamlit as st
import pdfplumber
import pandas as pd
import re
import io
from openpyxl.styles import Font, Alignment, PatternFill, Border
from openpyxl.utils import get_column_letter

# --- PAGE CONFIG ---
st.set_page_config(
    page_title="Koranizer — Bank Statement Converter",
    page_icon="📊",
    layout="wide"
)

# --- SVG ICONS ---
SVG_BANK = '<svg xmlns="http://www.w3.org/2000/svg" width="28" height="28" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.5" stroke-linecap="round" stroke-linejoin="round"><path d="M3 21h18"/><path d="M3 10h18"/><path d="M5 6l7-3 7 3"/><path d="M4 10v11"/><path d="M20 10v11"/><path d="M8 14v3"/><path d="M12 14v3"/><path d="M16 14v3"/></svg>'

SVG_UPLOAD = '<svg xmlns="http://www.w3.org/2000/svg" width="40" height="40" viewBox="0 0 24 24" fill="none" stroke="#7C9AB6" stroke-width="1.5" stroke-linecap="round" stroke-linejoin="round"><path d="M21 15v4a2 2 0 0 1-2 2H5a2 2 0 0 1-2-2v-4"/><polyline points="17 8 12 3 7 8"/><line x1="12" y1="3" x2="12" y2="15"/></svg>'

SVG_FILE = '<svg xmlns="http://www.w3.org/2000/svg" width="18" height="18" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.5" stroke-linecap="round" stroke-linejoin="round"><path d="M14 2H6a2 2 0 0 0-2 2v16a2 2 0 0 0 2 2h12a2 2 0 0 0 2-2V8z"/><polyline points="14 2 14 8 20 8"/><line x1="16" y1="13" x2="8" y2="13"/><line x1="16" y1="17" x2="8" y2="17"/></svg>'

SVG_DOWNLOAD = '<svg xmlns="http://www.w3.org/2000/svg" width="16" height="16" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M21 15v4a2 2 0 0 1-2 2H5a2 2 0 0 1-2-2v-4"/><polyline points="7 10 12 15 17 10"/><line x1="12" y1="15" x2="12" y2="3"/></svg>'

SVG_TABLE = '<svg xmlns="http://www.w3.org/2000/svg" width="18" height="18" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.5" stroke-linecap="round" stroke-linejoin="round"><rect x="3" y="3" width="18" height="18" rx="2" ry="2"/><line x1="3" y1="9" x2="21" y2="9"/><line x1="3" y1="15" x2="21" y2="15"/><line x1="9" y1="3" x2="9" y2="21"/></svg>'

# --- CUSTOM CSS ---
st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');
    
    * { font-family: 'Inter', -apple-system, sans-serif; }
    
    .block-container { padding-top: 2rem; max-width: 1100px; }
    
    /* Header */
    .app-header {
        display: flex;
        align-items: center;
        gap: 12px;
        padding: 0 0 0.5rem 0;
        border-bottom: 1px solid #E2E8F0;
        margin-bottom: 2rem;
    }
    .app-header .logo {
        background: #1A3A5C;
        color: white;
        width: 44px;
        height: 44px;
        border-radius: 10px;
        display: flex;
        align-items: center;
        justify-content: center;
    }
    .app-header .title {
        font-size: 1.35rem;
        font-weight: 600;
        color: #1A2332;
        letter-spacing: -0.02em;
    }
    .app-header .subtitle {
        font-size: 0.82rem;
        color: #64748B;
        margin-top: 1px;
    }
    
    /* Upload zone */
    .upload-zone {
        border: 1.5px dashed #CBD5E1;
        border-radius: 12px;
        padding: 2.5rem 2rem;
        text-align: center;
        background: #FAFBFC;
        margin-bottom: 1.5rem;
        transition: border-color 0.2s;
    }
    .upload-zone:hover { border-color: #94A3B8; }
    .upload-zone .label { font-size: 0.95rem; color: #475569; margin-top: 0.8rem; font-weight: 500; }
    .upload-zone .hint { font-size: 0.78rem; color: #94A3B8; margin-top: 0.3rem; }
    
    /* Metric cards */
    .metrics-row {
        display: grid;
        grid-template-columns: repeat(4, 1fr);
        gap: 12px;
        margin: 1.2rem 0;
    }
    .metric-card {
        background: #FAFBFC;
        border: 1px solid #E2E8F0;
        border-radius: 10px;
        padding: 1rem 1.2rem;
    }
    .metric-card .val {
        font-size: 1.15rem;
        font-weight: 600;
        color: #1A2332;
        font-variant-numeric: tabular-nums;
    }
    .metric-card .val.debit { color: #DC2626; }
    .metric-card .val.credit { color: #16A34A; }
    .metric-card .lbl {
        font-size: 0.72rem;
        color: #94A3B8;
        text-transform: uppercase;
        letter-spacing: 0.05em;
        margin-bottom: 4px;
        font-weight: 500;
    }
    
    /* Account info */
    .acct-info {
        display: flex;
        gap: 2rem;
        padding: 0.8rem 0;
        font-size: 0.88rem;
        color: #475569;
    }
    .acct-info strong { color: #1A2332; font-weight: 600; }
    
    /* Section header */
    .section-hdr {
        display: flex;
        align-items: center;
        gap: 8px;
        font-size: 0.9rem;
        font-weight: 600;
        color: #334155;
        margin: 1.2rem 0 0.6rem 0;
    }
    .section-hdr svg { color: #64748B; }
    
    /* Download btn */
    .stDownloadButton > button {
        background: #1A3A5C !important;
        color: white !important;
        border: none !important;
        border-radius: 8px !important;
        padding: 0.55rem 1.5rem !important;
        font-weight: 500 !important;
        font-size: 0.88rem !important;
        letter-spacing: 0.01em !important;
        transition: background 0.2s !important;
    }
    .stDownloadButton > button:hover {
        background: #0F2A42 !important;
    }
    
    /* Footer */
    .app-footer {
        border-top: 1px solid #E2E8F0;
        padding: 1rem 0 0.5rem 0;
        margin-top: 2rem;
        text-align: center;
        font-size: 0.75rem;
        color: #94A3B8;
    }
    
    /* Table tweaks */
    [data-testid="stDataFrame"] { border-radius: 8px; overflow: hidden; }
    
    /* Hide streamlit menu */
    #MainMenu, footer, header { visibility: hidden; }
</style>
""", unsafe_allow_html=True)


# --- CORE FUNCTIONS ---
def extract_header_info(pdf):
    page = pdf.pages[0]
    text = page.extract_text()
    match = re.search(r'(\d{13,})\s+[A-Z]{3}\s+(.*)', text)
    if match:
        return match.group(2).strip(), match.group(1)
    return "NAMA_PEMILIK", "NOMOR_REKENING"


def clean_cell(value):
    if value is None:
        return ""
    text = str(value)
    text = re.sub(r'(\d{2}:\d{2}:)\s*\n\s*(\d{2})', r'\1\2', text)
    text = re.sub(r'\s+', ' ', text.replace('\n', ' ')).strip()
    return text


def process_pdf(uploaded_file):
    with pdfplumber.open(uploaded_file) as pdf:
        owner_name, account_no = extract_header_info(pdf)
        
        all_rows = []
        for page in pdf.pages:
            tables = page.extract_tables()
            for table in tables:
                for row in table:
                    all_rows.append(row)
    
    if not all_rows:
        return None, None, None
    
    header_raw = all_rows[0]
    transactions = []
    for row in all_rows[1:]:
        if row == header_raw:
            continue
        cleaned = [clean_cell(c) for c in row]
        if len(cleaned) >= 6:
            debit = cleaned[3].replace(',', '') if cleaned[3] else "0.00"
            credit = cleaned[4].replace(',', '') if cleaned[4] else "0.00"
            balance = cleaned[5].replace(',', '') if cleaned[5] else "0.00"
            transactions.append({
                "Posting Date": cleaned[0],
                "Remark": cleaned[1],
                "Reference No": cleaned[2] if cleaned[2] else "-",
                "Debit": float(debit) if debit else 0.0,
                "Credit": float(credit) if credit else 0.0,
                "Balance": float(balance) if balance else 0.0,
            })
    
    return owner_name, account_no, transactions


def create_styled_excel(owner, acc_no, df):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Rekening Koran', index=False, startrow=4)
        ws = writer.sheets['Rekening Koran']
        
        hdr_fill = PatternFill(start_color="85B1DB", end_color="85B1DB", fill_type="solid")
        hdr_font = Font(color="FFFFFF", bold=False, size=11)
        alt_fill = PatternFill(start_color="F2F7FC", end_color="F2F7FC", fill_type="solid")
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left = Alignment(horizontal="left", vertical="center", wrap_text=True)
        right = Alignment(horizontal="right", vertical="center")
        
        ws.merge_cells('A1:F1')
        ws['A1'] = "LAPORAN REKENING KORAN"
        ws['A1'].font = Font(bold=True, size=14, color="1F4E78")
        ws['A2'], ws['B2'] = "Nama Pemilik:", owner
        ws['A3'], ws['B3'] = "No. Rekening:", f"'{acc_no}"
        for r in [2, 3]:
            ws.cell(row=r, column=1).font = Font(bold=True, size=11)
            ws.cell(row=r, column=2).font = Font(bold=True, size=11)
        
        h = 5
        for c, t in enumerate(df.columns, 1):
            cell = ws.cell(row=h, column=c)
            cell.value, cell.fill, cell.font = t, hdr_fill, hdr_font
            cell.alignment = Alignment(horizontal="left" if c <= 2 else "center", vertical="center", wrap_text=True)
        
        for rn, rd in enumerate(df.values, h + 1):
            for cn, v in enumerate(rd, 1):
                cell = ws.cell(row=rn, column=cn)
                if cn in [4, 5, 6]:
                    try:
                        cell.value = float(v)
                        cell.number_format = '#,##0.00'
                        cell.alignment = right
                    except: cell.value = v
                else:
                    cell.value = v
                    cell.alignment = center if cn in [1, 3] else left
                if rn % 2 == 0:
                    cell.fill = alt_fill
        
        for c, w in {1: 24, 2: 55, 3: 18, 4: 20, 5: 20, 6: 22}.items():
            ws.column_dimensions[get_column_letter(c)].width = w
    
    output.seek(0)
    return output


# --- HEADER ---
st.markdown(f"""
<div class="app-header">
    <div class="logo">{SVG_BANK}</div>
    <div>
        <div class="title">Koranizer MANDIRI</div>
        <div class="subtitle">Bank Statement to Excel Converter</div>
    </div>
</div>
""", unsafe_allow_html=True)

# --- UPLOAD ---
uploaded_files = st.file_uploader(
    "Upload Rekening Koran PDF",
    type=["pdf"],
    accept_multiple_files=True,
    label_visibility="collapsed"
)

if not uploaded_files:
    st.markdown(f"""
    <div class="upload-zone">
        {SVG_UPLOAD}
        <div class="label">Upload file PDF Rekening Koran</div>
        <div class="hint">Drag & drop atau klik Browse — mendukung batch upload</div>
    </div>
    """, unsafe_allow_html=True)

# --- RESULTS ---
if uploaded_files:
    for uploaded_file in uploaded_files:
        st.markdown("---")
        
        st.markdown(f"""
        <div class="section-hdr">
            {SVG_FILE} {uploaded_file.name}
        </div>
        """, unsafe_allow_html=True)
        
        with st.spinner("Memproses..."):
            owner, acc_no, transactions = process_pdf(uploaded_file)
        
        if transactions:
            df = pd.DataFrame(transactions)
            total_deb = df["Debit"].sum()
            total_cred = df["Credit"].sum()
            last_bal = df["Balance"].iloc[-1]
            
            # Account info
            st.markdown(f"""
            <div class="acct-info">
                <span><strong>Nama:</strong> {owner}</span>
                <span><strong>No. Rekening:</strong> {acc_no}</span>
            </div>
            """, unsafe_allow_html=True)
            
            # Metrics
            st.markdown(f"""
            <div class="metrics-row">
                <div class="metric-card">
                    <div class="lbl">Transaksi</div>
                    <div class="val">{len(transactions)}</div>
                </div>
                <div class="metric-card">
                    <div class="lbl">Total Debit</div>
                    <div class="val debit">Rp {total_deb:,.0f}</div>
                </div>
                <div class="metric-card">
                    <div class="lbl">Total Credit</div>
                    <div class="val credit">Rp {total_cred:,.0f}</div>
                </div>
                <div class="metric-card">
                    <div class="lbl">Saldo Akhir</div>
                    <div class="val">Rp {last_bal:,.0f}</div>
                </div>
            </div>
            """, unsafe_allow_html=True)
            
            # Table
            st.markdown(f'<div class="section-hdr">{SVG_TABLE} Data Transaksi</div>', unsafe_allow_html=True)
            
            disp = df.copy()
            disp["Debit"] = disp["Debit"].apply(lambda x: f"{x:,.2f}")
            disp["Credit"] = disp["Credit"].apply(lambda x: f"{x:,.2f}")
            disp["Balance"] = disp["Balance"].apply(lambda x: f"{x:,.2f}")
            
            st.dataframe(disp, use_container_width=True, height=420)
            
            # Download
            excel_data = create_styled_excel(owner, acc_no, df)
            xlsx_name = uploaded_file.name.replace(".pdf", ".xlsx")
            
            st.download_button(
                label=f"Download {xlsx_name}",
                data=excel_data,
                file_name=xlsx_name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key=f"dl_{uploaded_file.name}"
            )
        else:
            st.error(f"Tidak dapat mengekstrak data dari {uploaded_file.name}")

# --- FOOTER ---
st.markdown("""
<div class="app-footer">
    Koranizer v2.0 · Built with pdfplumber & Streamlit
</div>
""", unsafe_allow_html=True)
